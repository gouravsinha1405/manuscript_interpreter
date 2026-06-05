from __future__ import annotations

import argparse
import ast
import json
import re
from pathlib import Path
from typing import Any

from indic_transliteration import sanscript
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from annotation_app import build_page_context, request_groq_json


BASE_DIR = Path(__file__).resolve().parent
SPLIT_LINE_RE = re.compile(r"Split:\s*(.*)")
DEVANAGARI_RE = re.compile(r"[\u0900-\u097F]")
MANUAL_REVIEW_HINTS = {
    "भ्रगणयायनम": "सुरक्षित पंक्ति-पाठ: गण; वैकल्पिक शब्द-स्तर पढ़त: गणाय नमः",
}
READ_ME_TITLE = "Read Me"
STATUS_OPTIONS = 'approved,reject,needs_review'
UNKNOWN_TEXT = "<unknown>"
EDITABLE_FILL = PatternFill("solid", fgColor="FFF2CC")
APPROVED_FILL = PatternFill("solid", fgColor="C6E0B4")
REJECT_FILL = PatternFill("solid", fgColor="F4CCCC")
NEEDS_REVIEW_FILL = PatternFill("solid", fgColor="FCE5CD")
QWEN_VICCHEDA_MODEL = "qwen/qwen3-32b"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export page 002 review data to an Excel workbook for SME verification."
    )
    parser.add_argument(
        "--output",
        default="annotations/page_002_first_iteration_sme_review.xlsx",
        help="Path to the output .xlsx workbook",
    )
    return parser.parse_args()


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def combined_output(result: dict[str, Any]) -> str:
    return "\n".join([str(result.get("stdout", "")), str(result.get("stderr", ""))])


def to_devanagari(text: str) -> str:
    if not text or DEVANAGARI_RE.search(text):
        return text
    try:
        return sanscript.transliterate(text, sanscript.SLP1, sanscript.DEVANAGARI)
    except Exception:
        return text


def parse_sandhi_splits(result: dict[str, Any]) -> list[str]:
    splits: list[str] = []
    seen: set[str] = set()
    for line in combined_output(result).splitlines():
        match = SPLIT_LINE_RE.search(line)
        if not match:
            continue
        payload = match.group(1).strip()
        if payload == "No Splits Found":
            continue
        try:
            parsed = ast.literal_eval(payload)
        except Exception:
            continue
        if not isinstance(parsed, list) or not all(isinstance(part, str) for part in parsed):
            continue
        rendered = " + ".join(to_devanagari(part) for part in parsed)
        if rendered in seen:
            continue
        seen.add(rendered)
        splits.append(rendered)
    return splits


def parse_sandhi_split_records(result: dict[str, Any]) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    seen: set[str] = set()
    for line in combined_output(result).splitlines():
        match = SPLIT_LINE_RE.search(line)
        if not match:
            continue
        payload = match.group(1).strip()
        if payload == "No Splits Found":
            continue
        try:
            parsed = ast.literal_eval(payload)
        except Exception:
            continue
        if not isinstance(parsed, list) or not all(isinstance(part, str) for part in parsed):
            continue
        split_text = " + ".join(to_devanagari(part) for part in parsed)
        if split_text in seen:
            continue
        seen.add(split_text)
        records.append({"split": split_text, "meaning": UNKNOWN_TEXT})
    return records


def top_line_candidate(line: dict[str, Any]) -> dict[str, Any]:
    payload = line.get("interpretations", {}) if isinstance(line.get("interpretations"), dict) else {}
    candidates = payload.get("candidates", []) if isinstance(payload, dict) else []
    if isinstance(candidates, list) and candidates and isinstance(candidates[0], dict):
        return candidates[0]
    return {}


def best_index(best_contextual: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {
        str(item.get("observed_token_id", "")): item
        for item in best_contextual
        if isinstance(item, dict) and str(item.get("observed_token_id", "")).strip()
    }


def probe_index(probe_results: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {
        str(item.get("token", "")): item
        for item in probe_results
        if isinstance(item, dict) and str(item.get("token", "")).strip()
    }


def token_index(token_candidates: list[dict[str, Any]]) -> dict[tuple[int, int], list[dict[str, Any]]]:
    index: dict[tuple[int, int], list[dict[str, Any]]] = {}
    for item in token_candidates:
        if not isinstance(item, dict):
            continue
        if str(item.get("source", "")) != "ocr_segment":
            continue
        line_index = item.get("line_index")
        segment_index = item.get("segment_index")
        if not isinstance(line_index, int) or not isinstance(segment_index, int):
            continue
        index.setdefault((line_index, segment_index), []).append(item)
    return index


def render_segment_map(segments: list[dict[str, Any]]) -> str:
    rows: list[str] = []
    for segment in segments:
        if not isinstance(segment, dict):
            continue
        raw_text = str(segment.get("raw_text", "")).strip()
        contextual_text = str(segment.get("contextual_text", "")).strip()
        rows.append(f"{raw_text} -> {contextual_text}")
    return "\n".join(rows)


def render_token_splits(tokens: list[dict[str, Any]], best_by_id: dict[str, dict[str, Any]]) -> str:
    parts: list[str] = []
    for token in tokens:
        observed = str(token.get("token", "")).strip()
        observed_id = str(token.get("id", "")).strip()
        best = best_by_id.get(observed_id, {})
        best_surface = str(best.get("best_surface", "")).strip() or observed
        parts.append(f"{observed} -> {best_surface}")
    return "\n".join(parts)


def style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    wrap = Alignment(vertical="top", wrap_text=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = wrap
    sheet.auto_filter.ref = sheet.dimensions
    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 60)


def add_sheet_review_ux(sheet, editable_columns: list[str], status_column: str) -> None:
    if sheet.max_row < 2:
        return

    for column in editable_columns:
        for row in range(2, sheet.max_row + 1):
            sheet[f"{column}{row}"].fill = EDITABLE_FILL

    validation = DataValidation(type="list", formula1=f'"{STATUS_OPTIONS}"', allow_blank=True)
    validation.prompt = "Select SME review status"
    validation.promptTitle = "SME Status"
    validation.error = "Choose one of: approved, reject, needs_review"
    validation.errorTitle = "Invalid SME Status"
    sheet.add_data_validation(validation)
    validation.add(f"{status_column}2:{status_column}{sheet.max_row}")

    status_range = f"{status_column}2:{status_column}{sheet.max_row}"
    sheet.conditional_formatting.add(
        status_range,
        FormulaRule(formula=[f'${status_column}2="approved"'], stopIfTrue=False, fill=APPROVED_FILL),
    )
    sheet.conditional_formatting.add(
        status_range,
        FormulaRule(formula=[f'${status_column}2="reject"'], stopIfTrue=False, fill=REJECT_FILL),
    )
    sheet.conditional_formatting.add(
        status_range,
        FormulaRule(formula=[f'${status_column}2="needs_review"'], stopIfTrue=False, fill=NEEDS_REVIEW_FILL),
    )


def artifact_paths(page: dict[str, Any]) -> tuple[Path, Path, Path]:
    artifacts = page.get("artifacts", {}) if isinstance(page, dict) else {}
    token_graph_dir = BASE_DIR / "phase2_input" / str(artifacts.get("token_graph_dir", "page_002_token_graph_akshara"))
    contextual_dir = BASE_DIR / "phase2_input" / str(artifacts.get("contextual_dir", "page_002_contextual_graph_akshara_quick"))
    parser_probe_path = BASE_DIR / "phase2_input" / "page_002_parser_probe" / "probe_results.json"
    return token_graph_dir, contextual_dir, parser_probe_path


def add_instruction_sheet(workbook: Workbook) -> None:
    instructions = workbook.active
    instructions.title = READ_ME_TITLE
    instructions.append(["Sheet", "Purpose"])
    instructions.append([READ_ME_TITLE, "Workbook purpose and quick instructions for SME review"])
    instructions.append(["Line Review", "One row per manuscript line with OCR, current machine reading, and top probable meaning"])
    instructions.append(["Token Review", "One row per OCR token with best reading, sandhi/DP split evidence, and SME verification columns"])
    instructions.append([])
    instructions.append(["How to use", "Please review the machine reading, fill SME columns, and mark uncertain or incorrect readings in comments."])
    instructions.append(["Important note", "For भ्रगणयायनम the workbook carries both the safe reading 'गण' and the alternate interpretive hint 'गणाय नमः'."])
    instructions.append(["Status values", "Use only: approved, reject, needs_review. The status cells have a dropdown and will color automatically."])


def create_line_sheet(workbook: Workbook):
    line_sheet = workbook.create_sheet("Line Review")
    line_sheet.append([
        "line_index",
        "raw_ocr_line",
        "preferred_machine_line",
        "top_normalized_sanskrit",
        "top_probable_meaning_hindi",
        "uncertainty_notes",
        "segment_map",
        "token_split_summary",
        "sme_verified_line_text",
        "sme_status",
        "sme_comments",
    ])
    return line_sheet


def create_token_sheet(workbook: Workbook):
    token_sheet = workbook.create_sheet("Token Review")
    token_sheet.append([
        "line_index",
        "segment_index",
        "raw_ocr_line",
        "observed_ocr_token",
        "current_best_reading",
        "best_source",
        "lexicon_gloss",
        "sandhi_viccheda_candidates",
        "dp_split_candidate",
        "dp_split_meaning",
        "curated_review_hint",
        "line_probable_meaning_hindi",
        "line_uncertainty",
        "sme_verified_reading",
        "sme_verified_meaning",
        "sme_status",
        "sme_comments",
    ])
    return token_sheet


def create_viccheda_sheet(workbook: Workbook):
    viccheda_sheet = workbook.create_sheet("Viccheda Options")
    viccheda_sheet.append([
        "line_index",
        "segment_index",
        "raw_ocr_line",
        "observed_ocr_token",
        "qwen/qwen3-32b based viccheda",
        "final interpretation in english",
        "confidence",
    ])
    return viccheda_sheet


def plus_separated_viccheda(text: object) -> str:
    rendered = unknown_text(text)
    if rendered == UNKNOWN_TEXT:
        return rendered

    normalized = re.sub(r"\s*\+\s*", " + ", rendered)
    normalized = re.sub(r"\s*[|/,;\-–—]\s*", " + ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    if " + " in normalized:
        return normalized

    parts = [part for part in normalized.split(" ") if part]
    return " + ".join(parts) if len(parts) > 1 else normalized


def numeric_confidence(value: object) -> str:
    text = str(value).strip() if value is not None else ""
    if not text:
        return "0.00"

    lowered = text.lower()
    word_map = {
        "low": 0.30,
        "medium": 0.60,
        "high": 0.85,
        "very high": 0.95,
        "very_low": 0.20,
        "very low": 0.20,
        "मध्यम": 0.60,
        "कम": 0.30,
        "उच्च": 0.85,
    }
    if lowered in word_map:
        return f"{word_map[lowered]:.2f}"

    match = re.search(r"\d+(?:\.\d+)?", text)
    if not match:
        return "0.00"

    number = float(match.group(0))
    if number > 1.0:
        number = number / 100.0
    number = max(0.0, min(number, 1.0))
    return f"{number:.2f}"


def qwen_viccheda_result(
    observed_token: str,
    raw_line: str,
    cache: dict[str, dict[str, str]],
) -> dict[str, str]:
    token = observed_token.strip()
    if token in cache:
        return cache[token]

    if not token:
        result = {
            "viccheda": UNKNOWN_TEXT,
            "interpretation": UNKNOWN_TEXT,
            "confidence": UNKNOWN_TEXT,
        }
        cache[token] = result
        return result

    print(f"Qwen viccheda -> {token}", flush=True)

    system_prompt = (
        "You are a Sanskrit manuscript interpretation assistant. "
        "You receive one noisy OCR token from a manuscript line. "
        "Produce an interpretive viccheda, not necessarily a strict grammatical split. "
        "Prefer useful semantic chunking over formal overgeneration. "
        "Return only JSON with keys: viccheda, interpretation_english, confidence."
    )
    user_prompt = (
        f"Observed OCR token: {token}\n"
        f"Raw OCR line: {raw_line}\n"
        "Task: give the best possible interpretive viccheda for the full observed OCR token. "
        "Then give a concise final interpretation in English. "
        "If uncertain, still provide the most useful draft and mark confidence clearly."
    )

    try:
        response = request_groq_json(system_prompt, user_prompt, QWEN_VICCHEDA_MODEL)
    except Exception:
        response = {}

    result = {
        "viccheda": plus_separated_viccheda(response.get("viccheda")),
        "interpretation": unknown_text(response.get("interpretation_english")),
        "confidence": numeric_confidence(response.get("confidence")),
    }
    cache[token] = result
    return result


def segment_tokens_for_line(
    line_index: int,
    comparison_segments: list[dict[str, Any]],
    tokens_by_segment: dict[tuple[int, int], list[dict[str, Any]]],
) -> list[dict[str, Any]]:
    segment_tokens: list[dict[str, Any]] = []
    for segment in comparison_segments:
        if not isinstance(segment, dict):
            continue
        segment_index = segment.get("segment_index")
        if not isinstance(segment_index, int):
            continue
        segment_tokens.extend(tokens_by_segment.get((line_index, segment_index), []))
    return segment_tokens


def line_uncertainty_text(top_candidate: dict[str, Any]) -> str:
    uncertainty = top_candidate.get("uncertainty_notes", "")
    if isinstance(uncertainty, list):
        return " | ".join(str(item) for item in uncertainty)
    return str(uncertainty)


def append_line_review_row(
    line_sheet,
    line: dict[str, Any],
    top_candidate: dict[str, Any],
    comparison_segments: list[dict[str, Any]],
    segment_tokens: list[dict[str, Any]],
    best_by_id: dict[str, dict[str, Any]],
) -> None:
    line_sheet.append([
        int(line.get("line_index", 0)),
        str(line.get("raw_text", "")),
        str(line.get("preferred_text", "")),
        str(top_candidate.get("normalized_sanskrit", "")),
        str(top_candidate.get("interpretation_hindi", "")),
        line_uncertainty_text(top_candidate),
        render_segment_map(comparison_segments),
        render_token_splits(segment_tokens, best_by_id),
        "",
        "",
        "",
    ])


def append_token_review_rows(
    token_sheet,
    line: dict[str, Any],
    comparison_segments: list[dict[str, Any]],
    tokens_by_segment: dict[tuple[int, int], list[dict[str, Any]]],
    best_by_id: dict[str, dict[str, Any]],
    probe_by_token: dict[str, dict[str, Any]],
    top_candidate: dict[str, Any],
) -> None:
    line_index = int(line.get("line_index", 0))
    line_meaning_hindi = str(top_candidate.get("interpretation_hindi", ""))
    line_uncertainty = line_uncertainty_text(top_candidate)
    for segment in comparison_segments:
        if not isinstance(segment, dict):
            continue
        segment_index = segment.get("segment_index")
        if not isinstance(segment_index, int):
            continue
        for token in tokens_by_segment.get((line_index, segment_index), []):
            token_sheet.append(
                token_review_row(
                    line=line,
                    line_index=line_index,
                    segment_index=segment_index,
                    token=token,
                    best_by_id=best_by_id,
                    probe_by_token=probe_by_token,
                    line_meaning_hindi=line_meaning_hindi,
                    line_uncertainty=line_uncertainty,
                )
            )


def token_review_row(
    line: dict[str, Any],
    line_index: int,
    segment_index: int,
    token: dict[str, Any],
    best_by_id: dict[str, dict[str, Any]],
    probe_by_token: dict[str, dict[str, Any]],
    line_meaning_hindi: str,
    line_uncertainty: str,
) -> list[str | int]:
    observed_token = str(token.get("token", "")).strip()
    observed_id = str(token.get("id", "")).strip()
    best = best_by_id.get(observed_id, {})
    probe = probe_by_token.get(observed_token, {})
    best_candidate = token.get("best_candidate", {}) if isinstance(token.get("best_candidate"), dict) else {}
    sandhi_candidates = parse_sandhi_splits(dict(probe.get("sandhi", {}))) if isinstance(probe, dict) else []
    dp_candidates = probe.get("dp_candidates", []) if isinstance(probe, dict) else []
    top_dp = dp_candidates[0] if isinstance(dp_candidates, list) and dp_candidates and isinstance(dp_candidates[0], dict) else {}
    return [
        line_index,
        segment_index,
        str(line.get("raw_text", "")),
        observed_token,
        str(best.get("best_surface", "")) or observed_token,
        str(best.get("source", "")) or str(token.get("source", "")),
        str(best_candidate.get("gloss", "")),
        " | ".join(sandhi_candidates),
        str(top_dp.get("split", "")),
        str(top_dp.get("interpretation", "")),
        MANUAL_REVIEW_HINTS.get(observed_token, ""),
        line_meaning_hindi,
        line_uncertainty,
        "",
        "",
        "",
        "",
    ]


def unknown_text(value: object) -> str:
    text = str(value).strip() if value is not None else ""
    return text or UNKNOWN_TEXT


def append_viccheda_rows(
    viccheda_sheet,
    line: dict[str, Any],
    comparison_segments: list[dict[str, Any]],
    tokens_by_segment: dict[tuple[int, int], list[dict[str, Any]]],
    qwen_cache: dict[str, dict[str, str]],
) -> None:
    line_index = int(line.get("line_index", 0))
    raw_line = str(line.get("raw_text", ""))
    for segment in comparison_segments:
        if not isinstance(segment, dict):
            continue
        segment_index = segment.get("segment_index")
        if not isinstance(segment_index, int):
            continue
        for token in tokens_by_segment.get((line_index, segment_index), []):
            observed_token = str(token.get("token", "")).strip()
            qwen_result = qwen_viccheda_result(observed_token, raw_line, qwen_cache)
            viccheda_sheet.append([
                line_index,
                segment_index,
                raw_line,
                observed_token or UNKNOWN_TEXT,
                qwen_result["viccheda"],
                qwen_result["interpretation"],
                qwen_result["confidence"],
            ])


def build_workbook() -> Workbook:
    page = build_page_context()
    token_graph_dir, contextual_dir, parser_probe_path = artifact_paths(page)
    token_candidates = load_json(token_graph_dir / "token_candidates.json")
    best_contextual = load_json(contextual_dir / "best_contextual_candidates.json")
    probe_results = load_json(parser_probe_path)

    best_by_id = best_index(best_contextual)
    probe_by_token = probe_index(probe_results)
    tokens_by_segment = token_index(token_candidates)
    qwen_cache: dict[str, dict[str, str]] = {}

    workbook = Workbook()
    instructions = workbook.active
    instructions.title = READ_ME_TITLE
    add_instruction_sheet(workbook)

    line_sheet = create_line_sheet(workbook)
    token_sheet = create_token_sheet(workbook)
    viccheda_sheet = create_viccheda_sheet(workbook)

    lines = page.get("lines", []) if isinstance(page, dict) else []
    for line in lines:
        if not isinstance(line, dict):
            continue
        line_index = int(line.get("line_index", 0))
        top_candidate = top_line_candidate(line)
        comparison_segments = line.get("comparison_segments", []) if isinstance(line.get("comparison_segments"), list) else []
        segment_tokens = segment_tokens_for_line(line_index, comparison_segments, tokens_by_segment)
        append_line_review_row(line_sheet, line, top_candidate, comparison_segments, segment_tokens, best_by_id)
        append_token_review_rows(
            token_sheet,
            line,
            comparison_segments,
            tokens_by_segment,
            best_by_id,
            probe_by_token,
            top_candidate,
        )
        append_viccheda_rows(
            viccheda_sheet,
            line,
            comparison_segments,
            tokens_by_segment,
            qwen_cache,
        )

    style_sheet(instructions)
    style_sheet(line_sheet)
    style_sheet(token_sheet)
    style_sheet(viccheda_sheet)
    add_sheet_review_ux(line_sheet, editable_columns=["I", "J", "K"], status_column="J")
    add_sheet_review_ux(token_sheet, editable_columns=["N", "O", "P", "Q"], status_column="P")
    return workbook


def main() -> None:
    args = parse_args()
    output_path = BASE_DIR / args.output
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook()
    workbook.save(output_path)
    print(f"Wrote SME review workbook to {output_path}")


if __name__ == "__main__":
    main()
